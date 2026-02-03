import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelWriter:
    def __init__(self, model):
        self.model = model

    def _apply_style(self, ws, title):
        blue_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        white_font = Font(color='FFFFFF', bold=True)
        
        ws.merge_cells('A1:E1')
        ws['A1'] = title
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        for cell in ws[2]:
            cell.fill = blue_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal='center')

    def save(self, filename: str):
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # 1. Summary Sheet
            summary_df = pd.DataFrame({
                'Metric': list(self.model.returns.keys()),
                'Value': list(self.model.returns.values())
            })
            summary_df.to_excel(writer, sheet_name='Summary', index=False, startrow=1)
            self._apply_style(writer.sheets['Summary'], f"LBO Summary: {self.model.fetcher.ticker_symbol}")
            
            # 2. Projections Sheet
            self.model.projections.to_excel(writer, sheet_name='Projections', index=False, startrow=1)
            self._apply_style(writer.sheets['Projections'], "5-Year Projections & Debt Paydown")
            
            # 3. Sensitivity Sheet
            entry_range = [8.0, 9.0, 10.0, 11.0, 12.0]
            exit_range = [8.0, 9.0, 10.0, 11.0, 12.0]
            sens_df = self.model.sensitivity_analysis(entry_range, exit_range)
            sens_df.to_excel(writer, sheet_name='Sensitivity', startrow=1)
            self._apply_style(writer.sheets['Sensitivity'], "IRR Sensitivity: Entry vs Exit Multiple")
